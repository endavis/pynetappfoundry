"""Generate client lock reports."""

from __future__ import annotations

from datetime import datetime
from typing import Any

import click
from openpyxl import Workbook

import pynetappfoundry.cache.ontap.protocols.locks.mapping  # noqa: F401 - register TypeMapping
from pynetappfoundry.cli.decorators import with_config
from pynetappfoundry.cli.utils import print_debug, print_error, print_info, print_success
from pynetappfoundry.clients.ontap.api import ONTAPAPIClient
from pynetappfoundry.core.config import Config
from pynetappfoundry.core.models import ClusterConfig
from pynetappfoundry.models.ontap.protocols.locks import OntapClientLock
from pynetappfoundry.query import QuerySet


@click.command()
@click.option(
    "--filter",
    "-f",
    "filter",
    help='JSON filter: \'{"bu":"Business", "env":"Prod"}\'',
)
@with_config("Locks report failed")
def locks(
    config: Config,
    clusters: dict[str, dict[str, Any]],
) -> None:
    """Generate client lock reports.

    Creates an Excel workbook with all client locks across
    matching clusters.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = config.output_dir / f"locks_{timestamp}.xlsx"
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    if default_sheet:
        wb.remove(default_sheet)

    for name, details in clusters.items():
        print_info(f"Gathering lock data for {name}...")
        _gather_lock_data(name, details, config, wb)

    wb.save(filename)
    print_success(f"Report saved to {filename}")


def _gather_lock_data(
    name: str,
    details: dict[str, Any],
    config: Config,
    wb: Workbook,
) -> None:
    """Gather lock data from a single cluster."""
    try:
        cluster_config = ClusterConfig(**details)
        client = ONTAPAPIClient(cluster=cluster_config, config=config)

        locks_data: list[OntapClientLock] = QuerySet(OntapClientLock, client).all()

        ws = wb.create_sheet(name)
        ws.append(["Volume", "Protocol", "Type", "Path", "Lock", "State", "IP Address"])

        print_debug(f"Locks found: {len(locks_data)}")

        for lock in locks_data:
            try:
                lock_type = lock.type_
                if lock_type == "share_level":
                    ws.append(
                        [
                            lock.volume_name or "Unknown",
                            lock.protocol or "Unknown",
                            lock_type,
                            lock.path or "Unknown",
                            lock.share_lock_mode,
                            lock.state or "Unknown",
                            lock.client_address or "Unknown",
                        ]
                    )
                elif lock_type == "op_lock":
                    ws.append(
                        [
                            lock.volume_name or "Unknown",
                            lock.protocol or "Unknown",
                            lock_type,
                            lock.path or "Unknown",
                            lock.oplock_level,
                            lock.state or "Unknown",
                            lock.client_address or "Unknown",
                        ]
                    )
                else:
                    print_debug(f"Unknown lock type: {lock_type}")
            except Exception as e:
                print_error(f"Lock error: {e}")

    except Exception as e:
        print_error(f"Could not gather lock data for {name}: {e}")
        ws = wb.create_sheet(name)
        ws.append(["Error", str(e)])
