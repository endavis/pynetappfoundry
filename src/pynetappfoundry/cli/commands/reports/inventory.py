"""Generate inventory spreadsheet report.

Creates an Excel workbook with a NetApps sheet containing one row per
cluster with identity, network, cloud, node, and licensing columns.
Matches the layout of the Storage Devices inventory spreadsheet.
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any

import click
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pynetappfoundry.cli.decorators import with_config
from pynetappfoundry.cli.utils import print_info, print_success
from pynetappfoundry.core.cluster_entry import ClusterEntry
from pynetappfoundry.core.config import Config
from pynetappfoundry.data.source import DataSource
from pynetappfoundry.models.ontap.cloud.metadata.model import CloudMetadata
from pynetappfoundry.models.ontap.cluster.licensing.licenses import OntapLicensePackageResponse
from pynetappfoundry.models.ontap.cluster.model import ClusterInfo
from pynetappfoundry.models.ontap.cluster.nodes.model import OntapNodeResponse

# Maps cloud provider region codes to human-readable site names.
_REGION_TO_SITE: dict[str, str] = {
    # Azure
    "centralus": "Central",
    "eastus": "East 1",
    "eastus2": "East 2",
    "northcentralus": "North Central",
    "southcentralus": "South Central",
    "westus": "West 1",
    "westus2": "West 2",
    "westus3": "West 3",
    "westeurope": "West Europe",
    "northeurope": "North Europe",
    "uksouth": "UK South",
    "ukwest": "UK West",
    # AWS
    "us-east-1": "East 1",
    "us-east-2": "East 2",
    "us-west-1": "West 1",
    "us-west-2": "West 2",
    "eu-west-1": "Ireland",
    "eu-west-2": "London",
    "eu-central-1": "Frankfurt",
    "ap-southeast-1": "Singapore",
    "ap-southeast-2": "Sydney",
    "ap-northeast-1": "Tokyo",
}

_HEADERS = [
    "Name",
    "Division",
    "Application/ Business Unit",
    "Environment",
    "Contacts",
    "Notes",
    "Site",
    "Version",
    "Cloud Location",
    "Type",
    "License Expiration",
    "Subscription Name",
    "Resource Group Name",
    "Subscription ID",
    "Cluster Management IP",
    "System Manager",
    "BlueXP (NetApp Console) Connector IP",
    "BlueXP (NetApp Console) Name",
    "System ID",
    "Deployment Server",
    "New OCUM IP",
    "VM SKU Type",
    "Node-1 Name",
    "Node1 IP",
    "Node 1 Serial",
    "Node-2 Name",
    "Node2 IP",
    "Node 2 Serial",
    "Application Mapping",
    "Storage Account",
    "VM Names",
    "Load-Balancers",
    "Remark",
]

# Column widths (approximate, in character units)
_COL_WIDTHS = [
    20,
    12,
    22,
    12,
    40,
    35,
    14,
    14,
    14,
    12,
    20,
    35,
    55,
    38,
    20,
    25,
    40,
    28,
    38,
    18,
    30,
    22,
    25,
    16,
    25,
    25,
    16,
    25,
    35,
    35,
    35,
    55,
    20,
]


def _node_mgmt_ip(node: OntapNodeResponse) -> str:
    """Return the management IP for a node, falling back to management_interfaces list."""
    ip = node.management_interface.ip.address
    if not ip:
        for iface in node.management_interfaces:
            if iface.ip.address:
                return iface.ip.address
    return ip or ""


def _license_expiration(licenses: list[OntapLicensePackageResponse]) -> str:
    """Return earliest non-empty expiry date across all license packages, or blank."""
    earliest: str = ""
    for pkg in licenses:
        for lic in pkg.licenses:
            exp = lic.expiry_time
            if exp and (not earliest or exp < earliest):
                earliest = exp
    return earliest


def _short_version(ontap_version: str) -> str:
    """Extract short version like '9.16.1P3' from full ONTAP version string."""
    match = re.search(r"(\d+\.\d+\.\d+\w*)", ontap_version)
    return match.group(1) if match else ontap_version


def _cloud_type(cluster_info: ClusterInfo) -> str:
    """Derive the cluster type string."""
    if not cluster_info.is_cloud:
        return "OnTap Select"
    return "CVO HA" if cluster_info.is_ha else "CVO"


class InventoryReport:
    """Builds the NetApp inventory Excel workbook."""

    def __init__(self, config: Config, clusters: dict[str, Any]) -> None:
        self.config = config
        self.clusters = clusters
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "NetApps"

        self._header_fill = PatternFill("solid", fgColor="1F4E79")
        self._header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        self._body_font = Font(name="Calibri", size=10)
        self._link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        self._wrap = Alignment(wrap_text=True, vertical="top")
        self._top = Alignment(vertical="top")

    def _write_headers(self) -> None:
        """Write the header row with formatting."""
        for col_idx, header in enumerate(_HEADERS, 1):
            cell = self.ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = self._top
        self.ws.row_dimensions[1].height = 30
        self.ws.freeze_panes = "A2"
        self.ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}1"

    def _set_col_widths(self) -> None:
        for col_idx, width in enumerate(_COL_WIDTHS, 1):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _collect_cloud_meta(
        self,
        entry: ClusterEntry,
        cluster_info: ClusterInfo | None,
    ) -> list[CloudMetadata]:
        """Load cloud metadata from cache for CVO clusters only."""
        if not cluster_info or not cluster_info.is_cloud:
            return []
        ontap = entry.ontap
        if not ontap or not ontap.cloud:
            return []
        return list(ontap.cloud)

    def _build_row(
        self,
        name: str,
        entry: ClusterEntry,
        cluster_info: ClusterInfo | None,
        nodes: list[OntapNodeResponse],
        licenses: list[OntapLicensePackageResponse],
        cloud_meta: list[CloudMetadata],
    ) -> list[Any]:
        """Build a single data row for one cluster."""
        cloud_by_node: dict[str, CloudMetadata] = {cm.node: cm for cm in cloud_meta}
        first_cloud = cloud_meta[0] if cloud_meta else None

        div = entry.get("div", "")
        bu = entry.get("bu", "")
        env = entry.get("env", "")
        contacts = entry.get("contacts", "")
        cloud_cfg = entry.get("cloud", "")

        region_code = first_cloud.region if first_cloud else ""
        site = _REGION_TO_SITE.get(region_code.lower(), region_code)

        version = _short_version(cluster_info.ontap_version) if cluster_info else ""

        cloud_location = (
            first_cloud.provider if first_cloud else (cloud_cfg.title() if cloud_cfg else "")
        )

        ctype = _cloud_type(cluster_info) if cluster_info else ""

        lic_exp = _license_expiration(licenses)

        # Azure subscription info — try cloud metadata first, then TOML config
        sub_name = ""
        rg_name = ""
        rg_link = ""
        if first_cloud:
            rg_name = first_cloud.resource_group_name
            rg_link = first_cloud.resource_group_link
        if not sub_name:
            azure_cfg = entry.get("azure", {})
            if isinstance(azure_cfg, dict):
                sub_name = azure_cfg.get("location", "")
                if not rg_name:
                    rg_name = azure_cfg.get("resource_group", "")
        sub_id = first_cloud.account_id if first_cloud else ""
        rg_cell: Any = (rg_name, rg_link) if rg_name and rg_link else rg_name

        mgmt_ip = entry.get("ip", "")
        sys_mgr: Any = (f"https://{mgmt_ip}", f"https://{mgmt_ip}") if mgmt_ip else ""

        search_terms = {
            "div": div,
            "bu": bu,
            "app": entry.get("app", ""),
            "env": env,
        }
        connector = self.config.find_closest("connectors", search_terms)
        connector_ip: Any = ""
        connector_name = ""
        if connector:
            ip_raw = str(connector.get("ip", "")).strip()
            url = f"https://{ip_raw}/" if ip_raw else ""
            connector_ip = (url, url) if url else ""
            conn_azure = connector.get("azure", {})
            connector_name = conn_azure.get("vmname", "") if isinstance(conn_azure, dict) else ""

        system_id = cluster_info.cluster_uuid if cluster_info else ""

        aiqum = self.config.find_closest("aiqums", search_terms)
        aiqum_ip: Any = ""
        if aiqum:
            ip_raw = str(aiqum.get("ip", "")).strip()
            url = f"https://{ip_raw}" if ip_raw else ""
            aiqum_ip = (url, url) if url else ""

        vm_sku = first_cloud.instance_type if first_cloud else ""

        n1 = nodes[0] if len(nodes) > 0 else None
        n2 = nodes[1] if len(nodes) > 1 else None

        n1_name = n1.name if n1 else "NA"
        n1_ip = _node_mgmt_ip(n1) if n1 else "NA"
        n1_serial = n1.serial_number if n1 else "NA"
        n2_name = n2.name if n2 else "NA"
        n2_ip = _node_mgmt_ip(n2) if n2 else "NA"
        n2_serial = n2.serial_number if n2 else "N/A"

        vm_names_parts: list[str] = []
        for node in nodes:
            cm = cloud_by_node.get(node.name)
            if cm and cm.vm_name:
                vm_names_parts.append(cm.vm_name)
        vm_names = ", ".join(vm_names_parts)

        return [
            name,  # Name
            div,  # Division
            bu,  # Application/BU
            env,  # Environment
            contacts,  # Contacts
            "",  # Notes (blank for now)
            site,  # Site
            version,  # Version
            cloud_location,  # Cloud Location
            ctype,  # Type
            lic_exp,  # License Expiration
            sub_name,  # Subscription Name
            rg_cell,  # Resource Group Name (link)
            sub_id,  # Subscription ID
            mgmt_ip,  # Cluster Management IP
            sys_mgr,  # System Manager (link)
            connector_ip,  # BlueXP Connector IP (link)
            connector_name,  # BlueXP Connector Name
            system_id,  # System ID
            "NA",  # Deployment Server
            aiqum_ip,  # New OCUM IP (link)
            vm_sku,  # VM SKU Type
            n1_name,  # Node-1 Name
            n1_ip,  # Node1 IP
            n1_serial,  # Node 1 Serial
            n2_name,  # Node-2 Name
            n2_ip,  # Node2 IP
            n2_serial,  # Node 2 Serial
            "",  # Application Mapping (blank for now)
            "",  # Storage Account (blank for now)
            vm_names,  # VM Names
            "",  # Load-Balancers (blank for now)
            "",  # Remark
        ]

    def build(self) -> None:
        """Build the full workbook."""
        self._write_headers()

        ds = DataSource(self.config)
        row_idx = 2
        for name, entry in self.clusters.items():
            cluster_info = ds.query(ClusterInfo, cluster=name, source="auto").first()
            nodes = list(ds.query(OntapNodeResponse, cluster=name, source="auto"))
            licenses = list(ds.query(OntapLicensePackageResponse, cluster=name, source="auto"))
            cloud_meta = self._collect_cloud_meta(entry, cluster_info)

            row_data = self._build_row(name, entry, cluster_info, nodes, licenses, cloud_meta)

            for col_idx, value in enumerate(row_data, 1):
                wrap_cols = {5, 6, 29, 30, 31, 32}
                if isinstance(value, tuple):
                    display, url = value
                    cell = self.ws.cell(row=row_idx, column=col_idx, value=display)
                    if url:
                        cell.hyperlink = url
                    cell.font = self._link_font if url else self._body_font
                    cell.alignment = self._wrap if col_idx in wrap_cols else self._top
                else:
                    cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = self._body_font
                    cell.alignment = self._wrap if col_idx in wrap_cols else self._top

            row_idx += 1

        self._set_col_widths()

    def save(self, path: str) -> None:
        """Save the workbook to disk."""
        self.wb.save(path)


@click.command()
@click.option(
    "--filter",
    "-f",
    "filter",
    help='JSON filter: \'{"bu":"TAA", "env":"Prod"}\'',
)
@with_config("Inventory report generation failed")
def inventory(
    config: Config,
    clusters: dict[str, Any],
) -> None:
    """Generate inventory spreadsheet of all NetApp clusters.

    Creates an Excel workbook (NetApps sheet) with one row per cluster
    containing identity, network, cloud, node, and licensing columns.
    Matches the Storage Devices inventory spreadsheet layout.
    """
    if not clusters:
        raise click.ClickException("No clusters found matching the filter criteria")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = config.output_dir / f"inventory_{timestamp}.xlsx"

    print_info(f"Building inventory report for {len(clusters)} cluster(s)...")

    report = InventoryReport(config, clusters)
    report.build()
    report.save(str(filename))

    print_success(f"Report saved to {filename}")
