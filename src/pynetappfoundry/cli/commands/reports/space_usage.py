"""Generate space usage reports."""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

import click
from openpyxl import Workbook

from pynetappfoundry.cli.decorators import with_config
from pynetappfoundry.cli.utils import print_info, print_success
from pynetappfoundry.core.config import Config


@click.command("space-usage")
@click.option(
    "--filter",
    "-f",
    "filter",
    help='JSON filter: \'{"bu":"Business", "env":"Prod"}\'',
)
@with_config("Space usage report failed")
def space_usage(
    config: Config,
    clusters: dict[str, dict[str, Any]],
) -> None:
    """Generate space usage reports.

    Creates an Excel workbook with storage space usage data
    for all matching clusters.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = config.output_dir / f"space_usage_{timestamp}.xlsx"
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    if default_sheet:
        wb.remove(default_sheet)

    for name, details in clusters.items():
        print_info(f"Gathering space data for {name}...")
        _gather_space_data(name, details, config, wb)

    wb.save(filename)
    print_success(f"Report saved to {filename}")


def _gather_space_data(
    name: str,
    details: dict[str, Any],
    config: Config,
    wb: Workbook,
) -> None:
    """Gather space usage data from a single cluster."""
    from netapp_ontap import HostConnection
    from netapp_ontap.resources import Aggregate, Volume

    user, enc = config.get_user("clusters", name)

    try:
        with HostConnection(
            details["ip"],
            username=user,
            password=enc,
            verify=False,
        ):
            ws = wb.create_sheet(name)
            ws.append(
                [
                    "Type",
                    "Name",
                    "Total Size",
                    "Used Size",
                    "Available Size",
                    "Percent Used",
                    "State",
                ]
            )

            # Get aggregates
            for aggr in Aggregate.get_collection(fields="*"):
                aggr_dict = aggr.to_dict()
                space = aggr_dict.get("space", {})
                ws.append(
                    [
                        "Aggregate",
                        aggr_dict.get("name", "Unknown"),
                        space.get("block_storage", {}).get("size", 0),
                        space.get("block_storage", {}).get("used", 0),
                        space.get("block_storage", {}).get("available", 0),
                        space.get("block_storage", {}).get("full_threshold_percent", 0),
                        aggr_dict.get("state", "Unknown"),
                    ]
                )

            # Get volumes
            for vol in Volume.get_collection(fields="*"):
                vol_dict = vol.to_dict()
                space = vol_dict.get("space", {})
                ws.append(
                    [
                        "Volume",
                        vol_dict.get("name", "Unknown"),
                        space.get("size", 0),
                        space.get("used", 0),
                        space.get("available", 0),
                        space.get("percent_used", 0),
                        vol_dict.get("state", "Unknown"),
                    ]
                )

    except Exception as e:
        logging.error(f"Could not gather space data for {name}: {e}")
        ws = wb.create_sheet(name)
        ws.append(["Error", str(e)])
