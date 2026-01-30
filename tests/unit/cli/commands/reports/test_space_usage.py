"""Tests for space usage report generation."""

from __future__ import annotations

from pathlib import Path
from typing import Any
from unittest.mock import MagicMock, patch

import pytest

from pynetappfoundry.cli.commands.reports.space_usage import (
    BYTES_PER_TIB,
    ExcelReportBuilder,
    SpaceUsageCollector,
    VolumeData,
    bytes_to_tib,
    detect_cluster_type,
)


class TestBytesToTib:
    """Tests for bytes_to_tib function."""

    def test_one_tib(self) -> None:
        """Test conversion of exactly 1 TiB."""
        result = bytes_to_tib(BYTES_PER_TIB)
        assert result == 1.0

    def test_zero_bytes(self) -> None:
        """Test conversion of 0 bytes."""
        result = bytes_to_tib(0)
        assert result == 0.0

    def test_fractional_tib(self) -> None:
        """Test conversion with fractional TiB result."""
        # 512 GiB = 0.5 TiB
        gib_512 = 512 * (1024**3)
        result = bytes_to_tib(gib_512)
        assert result == 0.5

    def test_precision(self) -> None:
        """Test that result has 7 decimal places precision."""
        # 1 byte should produce a very small number
        result = bytes_to_tib(1)
        # 1 / (1024^4) = 9.094947e-13
        assert result == round(1 / BYTES_PER_TIB, 7)


class TestDetectClusterType:
    """Tests for detect_cluster_type function."""

    def test_single_node_cvo(self) -> None:
        """Test single node cluster returns CVO."""
        nodes = [{"name": "node1", "ha": {"enabled": False}}]
        result = detect_cluster_type(nodes)
        assert result == "CVO"

    def test_ha_cluster(self) -> None:
        """Test HA cluster returns CVO HA."""
        nodes = [
            {"name": "node1", "ha": {"enabled": True}},
            {"name": "node2", "ha": {"enabled": True}},
        ]
        result = detect_cluster_type(nodes)
        assert result == "CVO HA"

    def test_two_nodes_ha_disabled(self) -> None:
        """Test two nodes with HA disabled returns CVO."""
        nodes = [
            {"name": "node1", "ha": {"enabled": False}},
            {"name": "node2", "ha": {"enabled": False}},
        ]
        result = detect_cluster_type(nodes)
        assert result == "CVO"

    def test_empty_nodes_list(self) -> None:
        """Test empty nodes list returns CVO."""
        nodes: list[dict[str, Any]] = []
        result = detect_cluster_type(nodes)
        assert result == "CVO"

    def test_missing_ha_field(self) -> None:
        """Test node without ha field returns CVO."""
        nodes = [{"name": "node1"}, {"name": "node2"}]
        result = detect_cluster_type(nodes)
        assert result == "CVO"


class TestVolumeData:
    """Tests for VolumeData dataclass."""

    def test_create_volume_data(self) -> None:
        """Test creating VolumeData instance."""
        vol = VolumeData(
            cluster_name="cluster1",
            volume_name="vol1",
            div="Engineering",
            bu="Platform",
            app="MyApp",
            env="Prod",
            subapp="",
            cloud="azure",
            region="eastus",
            cluster_type="CVO HA",
            data_type="RW",
            size_tib=1.0,
            used_tib=0.5,
            available_tib=0.5,
            percent_used=50.0,
        )
        assert vol.cluster_name == "cluster1"
        assert vol.data_type == "RW"
        assert vol.percent_used == 50.0


class TestSpaceUsageCollector:
    """Tests for SpaceUsageCollector class."""

    def test_add_volume(self) -> None:
        """Test adding a volume to collector."""
        collector = SpaceUsageCollector()
        vol = VolumeData(
            cluster_name="cluster1",
            volume_name="vol1",
            div="Div",
            bu="BU",
            app="App",
            env="Prod",
            subapp="",
            cloud="azure",
            region="eastus",
            cluster_type="CVO",
            data_type="RW",
            size_tib=1.0,
            used_tib=0.5,
            available_tib=0.5,
            percent_used=50.0,
        )
        collector.add_volume(vol)
        assert len(collector.volumes) == 1
        assert vol in collector.volumes

    def test_hierarchy_tracking(self) -> None:
        """Test that hierarchy keys are tracked."""
        collector = SpaceUsageCollector()
        vol = VolumeData(
            cluster_name="cluster1",
            volume_name="vol1",
            div="Div",
            bu="BU",
            app="App",
            env="Prod",
            subapp="",
            cloud="azure",
            region="eastus",
            cluster_type="CVO",
            data_type="RW",
            size_tib=1.0,
            used_tib=0.5,
            available_tib=0.5,
            percent_used=50.0,
        )
        collector.add_volume(vol)
        expected_key = ("Div", "BU", "App", "Prod", "", "azure", "eastus", "CVO", "RW")
        assert expected_key in collector.hierarchy_keys


class TestExcelReportBuilder:
    """Tests for ExcelReportBuilder class."""

    @pytest.fixture
    def sample_collector(self) -> SpaceUsageCollector:
        """Create a sample collector with test data."""
        collector = SpaceUsageCollector()

        # Add some test volumes
        volumes = [
            VolumeData(
                cluster_name="azure-cluster-1",
                volume_name="vol1",
                div="Engineering",
                bu="Platform",
                app="MyApp",
                env="Prod",
                subapp="",
                cloud="azure",
                region="eastus",
                cluster_type="CVO HA",
                data_type="RW",
                size_tib=10.0,
                used_tib=5.0,
                available_tib=5.0,
                percent_used=50.0,
            ),
            VolumeData(
                cluster_name="azure-cluster-1",
                volume_name="vol2",
                div="Engineering",
                bu="Platform",
                app="MyApp",
                env="Prod",
                subapp="",
                cloud="azure",
                region="eastus",
                cluster_type="CVO HA",
                data_type="DP",
                size_tib=5.0,
                used_tib=4.5,
                available_tib=0.5,
                percent_used=90.0,
            ),
            VolumeData(
                cluster_name="aws-cluster-1",
                volume_name="vol3",
                div="Finance",
                bu="Treasury",
                app="FinApp",
                env="Dev",
                subapp="",
                cloud="aws",
                region="us-east-1",
                cluster_type="CVO",
                data_type="RW",
                size_tib=20.0,
                used_tib=19.0,
                available_tib=1.0,
                percent_used=95.0,
            ),
        ]
        for vol in volumes:
            collector.add_volume(vol)

        return collector

    def test_workbook_creation(self, tmp_path: Path) -> None:
        """Test that workbook is created successfully."""
        filename = tmp_path / "test.xlsx"
        builder = ExcelReportBuilder(str(filename))
        builder.close()
        assert filename.exists()

    def test_creates_three_sheets(
        self, tmp_path: Path, sample_collector: SpaceUsageCollector
    ) -> None:
        """Test that three worksheets are created."""
        import openpyxl

        filename = tmp_path / "test.xlsx"
        builder = ExcelReportBuilder(str(filename))
        builder.create_totals_sheet(sample_collector)
        builder.create_usage_breakdown_sheet(sample_collector)
        builder.create_volumes_sheet(sample_collector)
        builder.close()

        # Open with openpyxl to verify structure
        wb = openpyxl.load_workbook(filename)
        sheet_names = wb.sheetnames
        assert "Totals" in sheet_names
        assert "Usage Breakdown" in sheet_names
        assert "Volumes" in sheet_names
        assert len(sheet_names) == 3

    def test_totals_sheet_headers(
        self, tmp_path: Path, sample_collector: SpaceUsageCollector
    ) -> None:
        """Test that Totals sheet has correct headers."""
        import openpyxl

        filename = tmp_path / "test.xlsx"
        builder = ExcelReportBuilder(str(filename))
        builder.create_totals_sheet(sample_collector)
        builder.close()

        wb = openpyxl.load_workbook(filename)
        sheet = wb["Totals"]
        headers = [cell.value for cell in sheet[1]]
        expected = [
            "Cloud",
            "Cluster Type",
            "Data Type",
            "Total Size (TiB)",
            "Used (TiB)",
            "Available (TiB)",
            "Volume Count",
        ]
        assert headers == expected

    def test_totals_sheet_aggregation(
        self, tmp_path: Path, sample_collector: SpaceUsageCollector
    ) -> None:
        """Test that Totals sheet correctly aggregates data."""
        import openpyxl

        filename = tmp_path / "test.xlsx"
        builder = ExcelReportBuilder(str(filename))
        builder.create_totals_sheet(sample_collector)
        builder.close()

        wb = openpyxl.load_workbook(filename)
        sheet = wb["Totals"]

        # Get data rows (skip header)
        data: dict[tuple[str, str, str], dict[str, Any]] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                key = (str(row[0]), str(row[1]), str(row[2]))
                data[key] = {
                    "size": row[3],
                    "used": row[4],
                    "available": row[5],
                    "count": row[6],
                }

        # Azure CVO HA RW: 1 volume, 10 TiB
        assert ("AZURE", "CVO HA", "RW") in data
        assert data[("AZURE", "CVO HA", "RW")]["count"] == 1

        # AWS CVO RW: 1 volume, 20 TiB
        assert ("AWS", "CVO", "RW") in data
        assert data[("AWS", "CVO", "RW")]["count"] == 1

    def test_volumes_sheet_has_all_volumes(
        self, tmp_path: Path, sample_collector: SpaceUsageCollector
    ) -> None:
        """Test that Volumes sheet contains all volumes."""
        import openpyxl

        filename = tmp_path / "test.xlsx"
        builder = ExcelReportBuilder(str(filename))
        builder.create_volumes_sheet(sample_collector)
        builder.close()

        wb = openpyxl.load_workbook(filename)
        sheet = wb["Volumes"]

        # Count data rows (excluding header)
        row_count = sum(1 for row in sheet.iter_rows(min_row=2, values_only=True) if row[0])
        assert row_count == 3

    def test_volumes_sheet_headers(
        self, tmp_path: Path, sample_collector: SpaceUsageCollector
    ) -> None:
        """Test that Volumes sheet has correct headers."""
        import openpyxl

        filename = tmp_path / "test.xlsx"
        builder = ExcelReportBuilder(str(filename))
        builder.create_volumes_sheet(sample_collector)
        builder.close()

        wb = openpyxl.load_workbook(filename)
        sheet = wb["Volumes"]
        headers = [cell.value for cell in sheet[1]]
        expected = [
            "Cluster",
            "Volume",
            "Division",
            "BU",
            "App",
            "Environment",
            "SubApp",
            "Cloud",
            "Region",
            "Cluster Type",
            "Data Type",
            "Size (TiB)",
            "Used (TiB)",
            "Available (TiB)",
            "% Used",
        ]
        assert headers == expected

    def test_usage_breakdown_sheet_headers(
        self, tmp_path: Path, sample_collector: SpaceUsageCollector
    ) -> None:
        """Test that Usage Breakdown sheet has correct headers."""
        import openpyxl

        filename = tmp_path / "test.xlsx"
        builder = ExcelReportBuilder(str(filename))
        builder.create_usage_breakdown_sheet(sample_collector)
        builder.close()

        wb = openpyxl.load_workbook(filename)
        sheet = wb["Usage Breakdown"]
        headers = [cell.value for cell in sheet[1]]
        expected = [
            "Division",
            "BU",
            "App",
            "Environment",
            "SubApp",
            "Cloud",
            "Region",
            "Cluster Type",
            "Data Type",
            "Total Size (TiB)",
            "Used (TiB)",
            "Available (TiB)",
            "Volume Count",
        ]
        assert headers == expected


class TestSpaceUsageCommand:
    """Tests for the space_usage click command."""

    @pytest.fixture
    def mock_config(self) -> MagicMock:
        """Create a mock Config object."""
        config = MagicMock()
        config.output_dir = Path("/tmp/test-output")
        config.get_user.return_value = ("admin", "password123")
        return config

    @pytest.fixture
    def sample_clusters(self) -> dict[str, dict[str, Any]]:
        """Sample cluster configuration."""
        return {
            "test-cluster-1": {
                "name": "test-cluster-1",
                "ip": "10.0.0.1",
                "div": "Engineering",
                "bu": "Platform",
                "app": "MyApp",
                "env": "Prod",
                "subapp": "",
                "cloud": "azure",
                "region": "eastus",
            },
        }

    @pytest.fixture
    def mock_node(self) -> MagicMock:
        """Create a mock ONTAP Node."""
        node = MagicMock()
        node.to_dict.return_value = {
            "name": "node1",
            "ha": {"enabled": True},
        }
        return node

    @pytest.fixture
    def mock_volume(self) -> MagicMock:
        """Create a mock ONTAP Volume."""
        volume = MagicMock()
        volume.to_dict.return_value = {
            "name": "data_vol1",
            "type": "rw",
            "space": {
                "size": BYTES_PER_TIB,  # 1 TiB
                "used": BYTES_PER_TIB // 2,  # 0.5 TiB
                "available": BYTES_PER_TIB // 2,
            },
        }
        return volume

    def test_gather_cluster_data(
        self,
        mock_config: MagicMock,
        sample_clusters: dict[str, dict[str, Any]],
        mock_node: MagicMock,
        mock_volume: MagicMock,
    ) -> None:
        """Test gathering data from a cluster."""
        from pynetappfoundry.cli.commands.reports.space_usage import _gather_cluster_data

        collector = SpaceUsageCollector()

        with patch("netapp_ontap.host_connection.HostConnection") as mock_conn:
            mock_conn.return_value.__enter__ = MagicMock(return_value=None)
            mock_conn.return_value.__exit__ = MagicMock(return_value=None)

            with patch("netapp_ontap.resources.Node") as mock_node_class:
                mock_node_class.get_collection.return_value = [mock_node, mock_node]

                with patch("netapp_ontap.resources.Volume") as mock_vol_class:
                    mock_vol_class.get_collection.return_value = [mock_volume]

                    _gather_cluster_data(
                        "test-cluster-1",
                        sample_clusters["test-cluster-1"],
                        mock_config,
                        collector,
                    )

        # Should have collected 1 volume
        assert len(collector.volumes) == 1
        vol = collector.volumes[0]
        assert vol.cluster_name == "test-cluster-1"
        assert vol.cluster_type == "CVO HA"  # 2 nodes with HA enabled
        assert vol.data_type == "RW"
        assert vol.size_tib == 1.0
        assert vol.percent_used == 50.0

    def test_skips_root_volumes(
        self,
        mock_config: MagicMock,
        sample_clusters: dict[str, dict[str, Any]],
        mock_node: MagicMock,
    ) -> None:
        """Test that root volumes are skipped."""
        from pynetappfoundry.cli.commands.reports.space_usage import _gather_cluster_data

        collector = SpaceUsageCollector()

        root_volume = MagicMock()
        root_volume.to_dict.return_value = {
            "name": "svm_root",
            "type": "rw",
            "space": {"size": 1000000, "used": 500000, "available": 500000},
        }

        vol0_volume = MagicMock()
        vol0_volume.to_dict.return_value = {
            "name": "vol0",
            "type": "rw",
            "space": {"size": 1000000, "used": 500000, "available": 500000},
        }

        with patch("netapp_ontap.host_connection.HostConnection") as mock_conn:
            mock_conn.return_value.__enter__ = MagicMock(return_value=None)
            mock_conn.return_value.__exit__ = MagicMock(return_value=None)

            with patch("netapp_ontap.resources.Node") as mock_node_class:
                mock_node_class.get_collection.return_value = [mock_node]

                with patch("netapp_ontap.resources.Volume") as mock_vol_class:
                    mock_vol_class.get_collection.return_value = [root_volume, vol0_volume]

                    _gather_cluster_data(
                        "test-cluster-1",
                        sample_clusters["test-cluster-1"],
                        mock_config,
                        collector,
                    )

        # Both root volumes should be skipped
        assert len(collector.volumes) == 0

    def test_skips_zero_size_volumes(
        self,
        mock_config: MagicMock,
        sample_clusters: dict[str, dict[str, Any]],
        mock_node: MagicMock,
    ) -> None:
        """Test that volumes with zero size are skipped."""
        from pynetappfoundry.cli.commands.reports.space_usage import _gather_cluster_data

        collector = SpaceUsageCollector()

        zero_size_volume = MagicMock()
        zero_size_volume.to_dict.return_value = {
            "name": "data_vol",
            "type": "rw",
            "space": {"size": 0, "used": 0, "available": 0},
        }

        with patch("netapp_ontap.host_connection.HostConnection") as mock_conn:
            mock_conn.return_value.__enter__ = MagicMock(return_value=None)
            mock_conn.return_value.__exit__ = MagicMock(return_value=None)

            with patch("netapp_ontap.resources.Node") as mock_node_class:
                mock_node_class.get_collection.return_value = [mock_node]

                with patch("netapp_ontap.resources.Volume") as mock_vol_class:
                    mock_vol_class.get_collection.return_value = [zero_size_volume]

                    _gather_cluster_data(
                        "test-cluster-1",
                        sample_clusters["test-cluster-1"],
                        mock_config,
                        collector,
                    )

        # Zero size volume should be skipped
        assert len(collector.volumes) == 0

    def test_handles_connection_error(
        self,
        mock_config: MagicMock,
        sample_clusters: dict[str, dict[str, Any]],
    ) -> None:
        """Test that connection errors are handled gracefully."""
        from pynetappfoundry.cli.commands.reports.space_usage import _gather_cluster_data

        collector = SpaceUsageCollector()

        with patch("netapp_ontap.host_connection.HostConnection") as mock_conn:
            mock_conn.return_value.__enter__ = MagicMock(side_effect=Exception("Connection failed"))

            with patch(
                "pynetappfoundry.cli.commands.reports.space_usage.print_error"
            ) as mock_print:
                _gather_cluster_data(
                    "test-cluster-1",
                    sample_clusters["test-cluster-1"],
                    mock_config,
                    collector,
                )

                # Should print error but not raise
                mock_print.assert_called()

        # No volumes should be collected
        assert len(collector.volumes) == 0
