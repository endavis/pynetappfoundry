"""Tests for reports locks command."""

from __future__ import annotations

from pathlib import Path
from typing import Any
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from pynetappfoundry.cli.commands.reports.locks import _gather_lock_data
from pynetappfoundry.models.ontap.protocols.locks.model import OntapClientLock


def _make_lock(
    *,
    volume_name: str = "vol1",
    protocol: str = "cifs",
    type_: str = "share_level",
    path: str = "/data/file.txt",
    share_lock_mode: str = "read",
    oplock_level: str = "",
    state: str = "granted",
    client_address: str = "192.168.1.10",
) -> OntapClientLock:
    """Create an OntapClientLock with common defaults."""
    return OntapClientLock(
        volume_name=volume_name,
        protocol=protocol,
        type_=type_,
        path=path,
        share_lock_mode=share_lock_mode,
        oplock_level=oplock_level,
        state=state,
        client_address=client_address,
    )


@pytest.fixture
def mock_config() -> MagicMock:
    """Create a mock Config object."""
    config = MagicMock()
    config.get_user.return_value = ("admin", "password123")
    config.output_dir = Path("/tmp/test_output")
    config.get_ontap_api_settings.return_value = MagicMock(
        base_api_path="/api",
        timeout=30.0,
    )
    config.get_schema_location.return_value = Path("/tmp/schema")
    return config


@pytest.fixture
def sample_details() -> dict[str, Any]:
    """Sample cluster connection details."""
    return {"ip": "10.0.0.1", "bu": "Platform", "env": "Prod"}


class TestGatherLockData:
    """Tests for _gather_lock_data."""

    @patch("pynetappfoundry.cli.commands.reports.locks.QuerySet")
    @patch("pynetappfoundry.cli.commands.reports.locks.ONTAPAPIClient")
    def test_creates_sheet(
        self,
        mock_client_cls: MagicMock,
        mock_qs_cls: MagicMock,
        mock_config: MagicMock,
        sample_details: dict[str, Any],
    ) -> None:
        """Test that a sheet is created for the cluster."""
        lock = _make_lock()
        mock_qs_cls.return_value.all.return_value = [lock]

        wb = Workbook()
        default = wb.active
        if default:
            wb.remove(default)

        _gather_lock_data("cluster1", sample_details, mock_config, wb)

        assert "cluster1" in wb.sheetnames
        ws = wb["cluster1"]
        # Header row
        assert ws.cell(row=1, column=1).value == "Volume"
        assert ws.cell(row=1, column=7).value == "IP Address"
        # Data row
        assert ws.cell(row=2, column=1).value == "vol1"

    @patch("pynetappfoundry.cli.commands.reports.locks.QuerySet")
    @patch("pynetappfoundry.cli.commands.reports.locks.ONTAPAPIClient")
    def test_share_level_lock_row(
        self,
        mock_client_cls: MagicMock,
        mock_qs_cls: MagicMock,
        mock_config: MagicMock,
        sample_details: dict[str, Any],
    ) -> None:
        """Test that share_level lock row has correct columns."""
        lock = _make_lock(
            type_="share_level",
            share_lock_mode="read_write",
        )
        mock_qs_cls.return_value.all.return_value = [lock]

        wb = Workbook()
        default = wb.active
        if default:
            wb.remove(default)

        _gather_lock_data("cluster1", sample_details, mock_config, wb)

        ws = wb["cluster1"]
        # Row 2 is data (row 1 is header)
        assert ws.cell(row=2, column=1).value == "vol1"
        assert ws.cell(row=2, column=2).value == "cifs"
        assert ws.cell(row=2, column=3).value == "share_level"
        assert ws.cell(row=2, column=4).value == "/data/file.txt"
        assert ws.cell(row=2, column=5).value == "read_write"
        assert ws.cell(row=2, column=6).value == "granted"
        assert ws.cell(row=2, column=7).value == "192.168.1.10"

    @patch("pynetappfoundry.cli.commands.reports.locks.QuerySet")
    @patch("pynetappfoundry.cli.commands.reports.locks.ONTAPAPIClient")
    def test_op_lock_row(
        self,
        mock_client_cls: MagicMock,
        mock_qs_cls: MagicMock,
        mock_config: MagicMock,
        sample_details: dict[str, Any],
    ) -> None:
        """Test that op_lock row has correct columns."""
        lock = _make_lock(
            type_="op_lock",
            share_lock_mode="",
            oplock_level="batch",
        )
        mock_qs_cls.return_value.all.return_value = [lock]

        wb = Workbook()
        default = wb.active
        if default:
            wb.remove(default)

        _gather_lock_data("cluster1", sample_details, mock_config, wb)

        ws = wb["cluster1"]
        assert ws.cell(row=2, column=3).value == "op_lock"
        assert ws.cell(row=2, column=5).value == "batch"

    @patch("pynetappfoundry.cli.commands.reports.locks.print_error")
    @patch("pynetappfoundry.cli.commands.reports.locks.ONTAPAPIClient")
    def test_error_handling_creates_error_sheet(
        self,
        mock_client_cls: MagicMock,
        mock_print_error: MagicMock,
        mock_config: MagicMock,
        sample_details: dict[str, Any],
    ) -> None:
        """Test that connection failure creates an error sheet."""
        mock_client_cls.side_effect = Exception("Connection refused")

        wb = Workbook()
        default = wb.active
        if default:
            wb.remove(default)

        _gather_lock_data("cluster1", sample_details, mock_config, wb)

        assert "cluster1" in wb.sheetnames
        ws = wb["cluster1"]
        assert ws.cell(row=1, column=1).value == "Error"
        assert "Connection refused" in ws.cell(row=1, column=2).value
        mock_print_error.assert_called_once()

    @patch("pynetappfoundry.cli.commands.reports.locks.QuerySet")
    @patch("pynetappfoundry.cli.commands.reports.locks.ONTAPAPIClient")
    def test_empty_locks_creates_header_only(
        self,
        mock_client_cls: MagicMock,
        mock_qs_cls: MagicMock,
        mock_config: MagicMock,
        sample_details: dict[str, Any],
    ) -> None:
        """Test that no locks creates a sheet with header only."""
        mock_qs_cls.return_value.all.return_value = []

        wb = Workbook()
        default = wb.active
        if default:
            wb.remove(default)

        _gather_lock_data("cluster1", sample_details, mock_config, wb)

        ws = wb["cluster1"]
        assert ws.cell(row=1, column=1).value == "Volume"
        assert ws.cell(row=2, column=1).value is None
